#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Timestamp: "2025-10-01 20:31:51 (ywatanabe)"
# File: /ssh:sp:/home/ywatanabe/proj/scitex_repo/src/scitex/io/_save_modules/_excel.py
# ----------------------------------------
from __future__ import annotations
import os

__FILE__ = __file__
__DIR__ = os.path.dirname(__FILE__)
# ----------------------------------------

"""
Excel saving functionality for scitex.io.save
"""

import numpy as np


def _generate_pval_variants():
    connectors = ["", "_", "-"]
    pval_bases = ["pvalue", "p"]
    adjusted_variants = ["", "adjusted", "adj"]

    pval_variants = []
    for base in pval_bases:
        for conn in connectors:
            for adj in adjusted_variants:
                if adj:
                    variant = f"{base}{conn}{adj}"
                else:
                    if base == "p":
                        variant = f"{base}{conn}value" if conn else base
                    else:
                        variant = base
                if variant not in pval_variants:
                    pval_variants.append(variant)

    pval_variants.extend(["padj", "p-val"])
    pval_variants = sorted(set(pval_variants))
    return pval_variants


def _is_statistical_results(df):
    """
    Check if DataFrame contains statistical test results.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to check

    Returns
    -------
    bool
        True if DataFrame appears to contain statistical test results
    """
    import pandas as pd

    if not isinstance(df, pd.DataFrame):
        return False

    # Check for characteristic statistical test fields
    PVAL_VARIANTS = _generate_pval_variants()
    return any(col in df.columns for col in PVAL_VARIANTS)


def _apply_stats_styling(df, spath):
    """Apply conditional formatting to statistical results in Excel.

    Parameters
    ----------
    df : pd.DataFrame
        Statistical results DataFrame
    spath : str
        Path to Excel file

    Notes
    -----
    Color coding for p-values:
    - Red background: p < 0.001 (***)
    - Orange background: p < 0.01 (**)
    - Yellow background: p < 0.05 (*)
    - Light gray: p >= 0.05 (ns)
    """
    import pandas as pd

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    wb = load_workbook(spath)
    ws = wb.active

    fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    fill_orange = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )
    fill_yellow = PatternFill(
        start_color="FFE66D", end_color="FFE66D", fill_type="solid"
    )
    fill_gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    font_bold = Font(bold=True)

    PVAL_VARIANTS = _generate_pval_variants()
    pval_cols = [col for col in df.columns if col in PVAL_VARIANTS]

    for pval_col in pval_cols:
        pvalue_col_idx = df.columns.get_loc(pval_col) + 1

        for row_idx, pvalue in enumerate(df[pval_col], start=2):
            cell = ws.cell(row=row_idx, column=pvalue_col_idx)
            if pd.notna(pvalue):
                if pvalue < 0.001:
                    cell.fill = fill_red
                    cell.font = font_bold
                elif pvalue < 0.01:
                    cell.fill = fill_orange
                    cell.font = font_bold
                elif pvalue < 0.05:
                    cell.fill = fill_yellow
                    cell.font = font_bold
                else:
                    cell.fill = fill_gray

            if pd.notna(pvalue) and pvalue < 0.05:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).font = font_bold

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    wb.save(spath)


def save_excel(obj, spath, style=True, **kwargs):
    """Handle Excel file saving with optional statistical formatting.

    Parameters
    ----------
    obj : pd.DataFrame, dict, list of dict, or np.ndarray
        Object to save as Excel file
    spath : str
        Path where Excel file will be saved
    style : bool, default True
        If True, automatically apply conditional formatting to statistical results
        (p-values colored by significance level). Set to False to disable styling.
    **kwargs
        Additional keyword arguments passed to pandas.DataFrame.to_excel()

    Raises
    ------
    ValueError
        If object type cannot be saved as Excel file

    Notes
    -----
    When saving statistical test results (DataFrames with 'pvalue' column),
    automatic conditional formatting is applied unless style=False:

    - Red background: p < 0.001 (***)
    - Orange background: p < 0.01 (**)
    - Yellow background: p < 0.05 (*)
    - Light gray: p >= 0.05 (ns)
    - Bold font for significant results (p < 0.05)

    Examples
    --------
    >>> import scitex as stx
    >>> results = stx.stats.test_ttest_ind(x, y)
    >>> stx.io.save(results, 'results.xlsx')  # Auto-styled
    >>> stx.io.save(results, 'results.xlsx', style=False)  # No styling
    """
    # Lazy import to avoid circular import issues
    import pandas as pd

    # Convert to DataFrame
    if isinstance(obj, pd.DataFrame):
        df = obj
    elif isinstance(obj, dict):
        df = pd.DataFrame(obj)
    elif isinstance(obj, list):
        # Handle list of dicts (common for multiple test results)
        df = pd.DataFrame(obj)
    elif isinstance(obj, np.ndarray):
        df = pd.DataFrame(obj)
    else:
        raise ValueError(f"Cannot save object of type {type(obj)} as Excel file")

    # Save to Excel
    df.to_excel(spath, index=False, **kwargs)

    # Apply styling if enabled and data contains statistical results
    if style and _is_statistical_results(df):
        _apply_stats_styling(df, spath)


# EOF
